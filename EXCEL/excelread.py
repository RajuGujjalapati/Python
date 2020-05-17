from openpyxl import load_workbook
myexcel=load_workbook(filename=r'C:\Users\New\OneDrive\Desktop\PYTHON\test.xlsx')
#mysheet=myexcel.active
allsheets=myexcel.sheetnames
print(allsheets)
mysheet=myexcel[allsheets[1]]#go to
mysheet.title="raju"
rows=mysheet.max_row#calculating no. of rows
cols=mysheet.max_column
print(rows)
print(cols)
n=[]
a=[]
s=[]
startpos=1
while(rows>0):
    namestr='A'+str(startpos)
    agestr='B'+str(startpos)
    salstr='C'+str(startpos)
    name=mysheet[namestr].value
    age=mysheet[agestr].value
    sal=mysheet[salstr].value
    n.append(name)
    a.append(age)
    s.append(sal)
    rows -=1
    startpos +=1
print("Names are=", n)
print("age are=", a)
print("salaries are=", s)
myexcel.save(r'C:\Users\New\OneDrive\Desktop\PYTHON\test.xlsx')


